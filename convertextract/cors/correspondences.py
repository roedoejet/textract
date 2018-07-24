"""
Parse spreadsheets for correspondence lists.
"""

import os
import importlib
from openpyxl import load_workbook 
from openpyxl.styles import Font, Fill, Color
from openpyxl.cell import Cell
from .. import exceptions
import re
import random


# _*_ coding: utf-8 _*_
class Correspondence():
    
    def __init__(self, language):
        # Load workbook, either from correspondence spreadsheets, or user loaded
        this_dir = os.path.dirname(os.path.abspath(__file__))
        if not isinstance(language, type(None)):
            if re.search(r'\.xlsx', language):
                wb = load_workbook(language)
                ws = wb.active
            else:
                languageXLSX = language + ".xlsx"
                file_name = os.path.join(this_dir, "correspondence_spreadsheets", languageXLSX)
                wb = load_workbook(file_name)
                ws = wb.active

            # Create wordlist
            cor_list = []

            # Loop through rows in worksheet, create if statements for different columns and append Cors to cor_list.
            for entry in ws:
                newCor = {"from": "", "to": "", "before": "", "after": ""}
                for col in entry:
                    if col.column == 'A':
                        value = col.value
                        if type(value) == float or int:
                            value = str(value)
                        newCor["from"] = value
                    if col.column == 'B':
                        value = col.value
                        if type(value) == float or int:
                            value = str(value)
                        newCor["to"] = value
                    if col.column == 'C':
                        if col.value is not None:
                            value = col.value
                            if type(value) == float or int:
                                value = str(value)
                            newCor["before"] = value
                    if col.column == 'D':
                        if col.value is not None:
                            value = col.value
                            if type(value) == float or int:
                                value = str(value)
                            newCor["after"] = value
                cor_list.append(newCor)

            # Add match pattern regular expression
            for cor in cor_list:
                cor["match_pattern"] = self.rule_to_regex(cor)

            
            # To prevent feeding
            for cor in cor_list:
                # if output exists as input for another cor
                if cor['to'] in [temp_cor['from'] for temp_cor in cor_list]:
                    # assign a random, unique character as a temporary value. this could be more efficient
                    random_char = chr(random.randrange(9632, 9727))
                    # make sure character is unique
                    if [temp_char for temp_char in cor_list if 'temp' in list(temp_char.keys())]:
                        while random_char in [temp_char['temp'] for temp_char in cor_list if 'temp' in list(temp_char.keys())]:
                            random_char = chr(random.randrange(9632, 9727))
                    cor['temp'] = random_char

            # preserve rule ordering with regex, then apply context free changes from largest to smallest
            # context_sensitive_rules = filter(lambda x: x["before"] != "" or x["after"] != "", cor_list)
            context_sensitive_rules = [x for x in cor_list if (x['before'] != '' or x['after'] != "")]
            context_free_rules = [x for x in cor_list if x['before'] == "" and x["after"] == ""]
            context_free_rules.sort(key=lambda x: len(x["from"]), reverse=True)
            cor_list = context_sensitive_rules + context_free_rules
            self.cor_list = cor_list
    
    def rule_to_regex(self, rule):
        if rule['before'] is not None:
            before = rule["before"]
        else:
            before = ''
        if rule['after'] is not None:
            after = rule["after"]
        else:
            after = ''
        fromMatch = rule["from"]
        try:
            ruleRX = re.compile(before + fromMatch + after)
        except:
            raise Exception('Your regex is malformed. Escape all regular expression special characters in your conversion table.')
        return ruleRX    

    def apply_rules(self, to_parse):
        for cor in self.cor_list:
            # print('first')
            if re.search(cor["match_pattern"], to_parse):
                # if a temporary value was assigned
                if 'temp' in list(cor.keys()):
                    # turn the original value into the temporary one
                    # print("turned orig value " + cor['from'] + " to temp " + cor["temp"])
                    to_parse = re.sub(cor["from"], cor["temp"], to_parse)
                else:
                    # else turn it into the final value
                    # print("turned orig value " + cor['from'] + " to final " + cor["to"])
                    # print(to_parse)
                    to_parse = re.sub(cor["from"], cor["to"], to_parse)
                    # print(to_parse)
        # print(to_parse)
        # transliterate temporary values
        for cor in self.cor_list:
            # print('second')
            # transliterate temp value to final value if it exists, otherwise pass
            try:
                if cor['temp'] and re.search(cor['temp'], to_parse):
                    # print("turned temp " + cor['temp'] + " to final " + cor["to"])
                    to_parse = re.sub(cor['temp'], cor['to'], to_parse)
            except KeyError:
                pass
        return to_parse

